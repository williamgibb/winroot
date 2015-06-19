from __future__ import print_function

'''
analyzer.py

library for handling root data from winrhizotron

this whole file should really be broken between a library file and a program

'''
import argparse
import logging
import os
import sys
# Third Party
import openpyxl
# Custom
import fields
import root
import tube
import utility

from errors import AnalyzerError, DataError, SerializationError


# logging config
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s: %(levelname)s: %(message)s [%(filename)s:%(funcName)s]')
log = logging.getLogger(__name__)



class Analyzer(object):
    def __init__(self,
                 additional_root_fields=None,
                 required_sheet_names=None):
        self.root_fields = fields.RootDataFields(additional_fields=additional_root_fields)
        self.synthesis_fields = fields.SynthesisDataFields()
        self.required_sheet_names = {'root_data':'ROOT',
                                     'synthesis_data': 'Synthesis',}

        if required_sheet_names:
            for key in ['root_data', 'synthesis_data']:
                if key not in required_sheet_names:
                    raise AnalyzerError('Missing key in required_sheet_names - [{}]'.format(key))
            self.required_sheet_names = required_sheet_names

        self.root_data = {} # Tube number -> list of roots from that tube.
        self.synthesis_data = {} # tube number -> rootidentity -> data
        self.tubes = [] # List of tube objects


    def insert(self, fp):
        log.info('Opening workbook [{}]'.format(fp))
        try:
            wb = openpyxl.load_workbook(filename=fp)
        except:
            log.exception('')
            raise AnalyzerError('Failed to load workbook [{}]'.format(fp))
        sheets = set(wb.get_sheet_names())
        if not sheets.issuperset(set(self.required_sheet_names.values())):
            raise AnalyzerError('Workbook is missing expected sheet names {}'.format(list(self.required_sheet_names)))
        self.parse(wb)

    def _process_synthesis_table(self, ws):
        log.info('Extracting synthesis data')
        synthesis_data = utility.build_data_from_fields(ws, self.synthesis_fields)
        for d in synthesis_data:
            tn = d.get('Tube#')
            if tn not in self.synthesis_data:
                self.synthesis_data[tn] = {}
            sd = self.synthesis_data.get(tn)
            root_identity = root.RootIdentity(rootname=d.get('RootName'),
                                              location=d.get('Location#'),
                                              birthsession=d.get('BirthSession'))
            if root_identity in sd:
                raise DataError('Duplicate root encountered in synthesis data: {}'.format(root_identity))
            sd[root_identity] = d
        return True

    def _process_root_table(self, ws):
        log.info('Extracting root table')
        root_data = utility.build_data_from_fields(ws, self.root_fields)
        log.info('Building roots from root_data')
        for d in root_data:
            tn = d.get('Tube#')
            # XXX Use collections.Defaultdict
            if tn not in self.root_data:
                self.root_data[tn] = []
            rl = self.root_data.get(tn)
            root_obj = self._root_from_dict(d)
            rl.append(root_obj)
        return True

    def _root_from_dict(self, d):

        attr_map = dict(self.root_fields.required_attributes)
        for k, v in self.synthesis_fields.required_attributes.items():
            if k in attr_map:
                continue
            attr_map[k] = v

        root_obj = root.Root(attr_map=attr_map,
                             rootname=d.get('RootName'),
                             location=d.get('Location#'),
                             birthsession=d.get('BirthSession'))

        # Check for anomalous roots
        num_tips = d.get('NumberOfTips')
        tip_liv_status = d.get('TipLivStatus')
        if num_tips == 1:
            root_obj.anomaly = False
            if tip_liv_status.startswith('A'):
                root_obj.isAlive = 'A'
            # XXX Configurable status!
            elif tip_liv_status.startswith(('D', 'G')):
                root_obj.isAlive = 'G'
            else:
                raise DataError('Unknown tip_liv_status [{}][{}]'.format(root_obj.identity, tip_liv_status))
        else:
            root_obj.anomaly = True
            root_obj.isAlive = 'A'
        # Set required attributes
        for k in self.root_fields.required_attributes:
            root_obj.set(k, d.get(k))
        # Check to see if the current root is gone
        # XXX Configurable value!
        if root_obj.isAlive.startswith(('D','G')):
            # XXX Eww hardcorded attribute access!
            root_obj.DeathSession = root_obj.get('Session#')
        # Now we add arbitrary keys to the root.
        for key in self.root_fields.custom_attributes:
            root_obj.set(key, d.get(key))
        return root_obj

    def parse(self, wb):
        log.debug('Parsing workbook')
        # First, we want to grab data from the synthesis table though.
        root_sheet = wb.get_sheet_by_name(self.required_sheet_names.get('root_data'))
        synthesis_sheet = wb.get_sheet_by_name(self.required_sheet_names.get('synthesis_data'))

        self._process_synthesis_table(ws=synthesis_sheet)
        self._process_root_table(ws=root_sheet)

        if set(self.root_data.keys()) != set(self.synthesis_data.keys()):
            log.error('# Root data keys [{}]'.format(len(self.root_data.keys())))
            log.error('# Syn  data keys [{}]'.format(len(self.synthesis_data.keys())))
            log.error(self.synthesis_data)
            raise DataError('Tube numbers from root_data does not match the tube numbers from the synthesis_data')

        log.info('Processing collected data')
        for tn in self.root_data:
            log.info('Processing data for tube [{}]'.format(tn))
            tube_obj = tube.Tube(tn)
            raw_roots = self.root_data.get(tn)

            for root_obj in raw_roots:
                rsession = root_obj.get('Session#')
                if rsession > tube_obj.maxSessionCount:
                    tube_obj.maxSessionCount = rsession
                    log.debug('Max session count updated to {}'.format(tube_obj.maxSessionCount))
                if rsession not in tube_obj.sessionDates:
                    tube_obj.sessionDates[rsession] = root_obj.get('Date')
                    log.debug('Inserted session {} - Date {}'.format(rsession, root_obj.get('Date')))
            final_roots = []
            log.info('Inserting roots into tube [{}]'.format(tn))
            for root_obj in raw_roots:
                tube_obj.insert_or_update_root(root_obj)
                if root_obj.get('Session#') == tube_obj.maxSessionCount:
                    final_roots.append(root_obj)
            log.info('Finalizing roots')
            for root_obj in final_roots:
                status = tube_obj.finalize_root(root_obj)
                if not status:
                    log.error('Failed to finalize root {}'.format(root_obj.identity))
            log.info('Inserting synthesis data')
            # Insert the sythesis data (containing the tip stats) into the roots.
            sdata = self.synthesis_data.get(tn)
            tube_obj.insert_synthesis_data(sdata)
            self.tubes.append(tube_obj)

    def write(self, fp):
        if not self.tubes:
            raise SerializationError('No tubes available to serialize data from')

        header = sorted(self.root_fields.identity_attributes.keys())
        header.extend(sorted([k for k in self.root_fields.required_attributes.keys() if k not in header]))
        header.extend(sorted([k for k in self.synthesis_fields.required_attributes.keys() if k not in header]))
        header.extend(sorted([k for k in root.Root.fixed_attributes if k not in header]))
        log.debug('Header row is {}'.format(header))


        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws.title = 'Compiled Data'  # XXX Custom title?
        row_index = 1
        for i, v in enumerate(header, 1):
            col = openpyxl.cell.get_column_letter(i)
            ws.cell('{x}{y}'.format(x=col, y=row_index)).value = v

        row_index = row_index + 1
        for tube_obj in self.tubes:
            log.info('Writing out data for tube [{}]'.format(tube_obj.tubeNumber))
            for root_obj in tube_obj:
                for i, v in enumerate(header, 1):
                    col = openpyxl.cell.get_column_letter(i)
                    if v in self.root_fields.required_attributes:
                        v = self.root_fields.required_attributes.get(v)
                    elif v in self.synthesis_fields.required_attributes:
                        v = self.synthesis_fields.required_attributes.get(v)
                    cv = getattr(root_obj, v, 'NO VALUE')
                    ws.cell('{x}{y}'.format(x=col, y=row_index)).value = cv
                row_index = row_index + 1

        wb.save(filename=fp)
        return True

#
# Main program functions
#


def main(options):
    #
    #   DOES NOT HANDLE EXCEPTION LISTS
    #

    # file path verification and options handling
    if not options.verbose:
        log.info('Output will not be verbose')
        logging.disable(logging.DEBUG)

    if not os.path.isfile(options.src_file):
        log.error('specified source is not a file')
        sys.exit(-1)

    if os.path.exists(options.output):
        log.warning('Specified output file already exists.\n')
        if not utility.query_yes_no('Do you want to overwrite that file?', 'no'):
            log.info('Exiting')
            sys.exit(-1)

    analyzer = Analyzer()
    analyzer.insert(options.src_file)
    analyzer.write(options.output)

    log.info('Done processing all data')
    sys.exit(0)


def root_options():
    parser = argparse.ArgumentParser(prog=__name__)
    parser.add_argument('-s', '--source', dest='src_file', required=True, type=str, action='store',
                        help='Source xlsx file to process')
    parser.add_argument('-o', '--output', dest='output', required=True, type=str, action='store',
                        help='Define the output file (xlsx)')
    parser.add_argument('-v', '--verbose', dest='verbose', default=False, action='store_true',
                        help='Enable verbose output')
    return parser


if __name__ == "__main__":
    p = root_options()
    opts = p.parse_args()

    main(opts)
