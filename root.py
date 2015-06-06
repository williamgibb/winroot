'''
root.py

library for handling root data from winrhizotron

this whole file should really be broken between a library file and a program

'''
import argparse
import logging
import os
import sys

import openpyxl


# logging config
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s: %(levelname)s: %(message)s [%(filename)s:%(funcName)s]')
log = logging.getLogger(__name__)


class Root(object):
    def __init__(self, rootName, location, birthSession):
        self.rootName = rootName
        self.location = location
        self.birthSession = birthSession
        self.identity = (rootName, location, birthSession)
        self.tipIdentity = (birthSession, location)
        self.tube = ''
        self.session = ''
        self.goneSession = ''
        self.anomaly = ''
        self.order = ''
        self.highestOrder = ''
        self.censored = ''
        self.aliveTipsAtBirth = ''
        self.aliveTipsAtGone = ''
        self.avgDiameter = ''
        self.isAlive = ''


class Tube(object):
    def __init__(self, tubeNumber):
        self.tubeNumber = tubeNumber
        self.roots = []
        self.maxSessionCount = 0
        self.tipStats = ''
        self.sessionDates = {}
        self.index = 0

    # this iter code is broken and i'm not certain why
    def __iter__(self):
        return self

    def next(self):
        if self.index == len(self.roots):
            self.index = 0
            raise StopIteration
        root = self.roots[self.index]
        self.index = self.index + 1
        return root

    def add_root(self, root):
        root.tube = self.tubeNumber
        self.roots.append(root)

    def insert_or_update_root(self, root):
        # insert root if the root identity is new
        insert = True
        for existingRoot in self:
            if root.identity == existingRoot.identity:
                insert = False
                # If there was a change, update the root attributes
                if existingRoot.isAlive.startswith('A') and root.isAlive.startswith('G'):
                    # root changed from A to G
                    logging.debug('Changing root from A to G')
                    existingRoot.goneSession = root.goneSession
                    existingRoot.isAlive = root.isAlive
                elif existingRoot.isAlive.startswith('G') and root.isAlive.startswith('A'):
                    # root changed from G to A
                    logging.debug('Changing root from G to A')
                    existingRoot.goneSession = ''
                    existingRoot.isAlive = root.isAlive
        # add the root to the tube
        if insert:
            # possible to insert a root at the last session.  likely rare though.
            # need to finalize this root before adding it into the tube.
            logging.debug('Adding root to tube %s' % (str(root.identity)))
            self.add_root(root)
        return True

    def finalize_root(self, root):
        finalized = False
        for existingRoot in self:
            if root.identity == existingRoot.identity:
                if root.session == self.maxSessionCount:
                    existingRoot.highestOrder = root.order
                if existingRoot.isAlive.startswith('A'):
                    existingRoot.goneSession = 0
                    existingRoot.censored = 1
                if existingRoot.isAlive.startswith('G'):
                    existingRoot.censored = 0
                finalized = True
        return finalized


def get_test_ws():
    wb = openpyxl.load_workbook(filename=r'unedited.xlsx')
    ws = wb.get_sheet_by_name('T008 Root Synth')
    return ws


''' general helper function for ws row/column access'''


def get_index_by_value(cell_list, value):
    ''' given a list of cells and a value, return the index in the list of
    that cell value.  this returns the index of the first time the value
    is found.
    
    if the value is not found, return false
    '''
    index = False
    for cell in cell_list:
        if cell.value == value:
            try:
                index = cell_list.index(cell)
            except ValueError, e:
                logging.error('Failed to get index for value after finding a match')
                logging.error('%s' % (e))
                return index
            break
    return index


def get_index_at_null(cell_list):
    ''' get the index value for the first NoneType object present in cell list'''
    index = False
    for cell in cell_list:
        if not cell.value:
            index = cell_list.index(cell)
            break
    return index


''' helper functions for finding the import values from the root data table'''


def build_root_data_table(ws, rootIndexData):
    ''' BUILD A LIST OF DATA REPRESENTING EACH ROW IN WORKSHEET UP 
    TO THE SYNTHESIS TABLE
    '''
    rootData = []
    # get index values for each important column
    tableHeader = ws.rows[0]
    for key in rootIndexData:
        index = get_index_by_value(tableHeader, key)
        if not index:
            logging.error('Failed to obtain header value: %s' % (key))
            return False
        rootIndexData[key] = index
    endOfRootData = get_index_at_null(ws.columns[0])
    if not endOfRootData:
        logging.error('Failed to find end of root data.')
        return False
    # this slice includes the header row in rootData.  This lets us
    # rebuild indices, since the dict iterator is not sorted
    for row in ws.rows[:endOfRootData]:
        rootRow = []
        for key in rootIndexData:
            value = row[rootIndexData[key]].value
            rootRow.append(value)
        rootData.append(rootRow)
    return rootData


def process_worksheet(ws, exceptions_list=False):
    '''Process a worksheet, and build a tube object containing all the root data.
    This data can then be dumped out when all tubes have been processed.'''
    rootIndexData = {
        'RootName': 0,
        'Location#': 0,
        'Session#': 0,
        'BirthSession': 0,
        'TotAvgDiam(mm/10)': 0,
        'TipLivStatus': 0,
        'RootNotes': 0,
        'NumberOfTips': 0,
        'Date': 0,
    }
    #
    # exception handling
    exceptions = None
    if exceptions_list:
        try:
            exceptions = exceptions_list[tubeNumber]
        except KeyError, e:
            logging.error('No exceptions list found for tube number: %s' % (str(tubeNumber)))
    # basic information
    tubeNumber = get_tube_number(ws)
    tube = Tube(tubeNumber)
    logging.debug('Processing Data for tube #: %s' % (str(tubeNumber)))
    #
    root_data = build_root_data_table(ws, rootIndexData)
    if not root_data:
        logging.error('Failed to get root data from sheet: %s' % (ws.title))
        return False
    headerRow = root_data[0]
    for key in rootIndexData:
        try:
            index = headerRow.index(key)
        except ValueError, e:
            logging.error('Failed to obtain header value: %s' % (key))
            logging.error('%s' % (e))
            return False
        rootIndexData[key] = index
    # this loop will not handle exceptional roots, it just builts a list of root 
    # object for each row of root data
    logging.debug('Building roots from root_data')
    maxSessionCount = 1
    sessionDates = {}
    roots = []
    for row in root_data[1:]:
        if exceptions:
            for i in exceptions:
                rootName = row[rootIndexData['RootName']]
                rootLocation = row[rootIndexData['Location#']]
                rootSession = row[rootIndexData['Sesssion#']]
                #
                # EXCEPTION ROOT HANDLING WOULD WOULD GO HERE
                #
                pass
        root = root_from_row(row, rootIndexData)
        roots.append(root)
        # process session statistics
        if root.session > maxSessionCount:
            maxSessionCount = root.session
            logging.debug('MaxSessionCount updated to %s' % (str(maxSessionCount)))
        if root.session not in sessionDates:
            sessionDates[root.session] = row[rootIndexData['Date']]
            logging.debug('Inserted session %s- Date %s ' % (str(root.session), sessionDates[root.session],))
    # no longer need to keep this data in memory
    del (root_data)
    # need to know what session value to use for finalizing root values.
    tube.maxSessionCount = maxSessionCount
    tube.sessionDates = sessionDates
    # add each root to the tube
    logging.debug('Inserting roots')
    finalRoots = []
    for root in roots:
        tube.insert_or_update_root(root)
        if root.session == maxSessionCount:
            finalRoots.append(root)
    # finalize roots
    logging.debug('Finalizing roots')
    for root in finalRoots:
        status = tube.finalize_root(root)
        if not status:
            logging.error('Failed to finalize root %s' % (str(root.identity)))
            print root.identity

    # need to calculate alive/dead tip numbers
    tipStats = tip_stats(tube)
    for r in tube.roots:
        r.aliveTipsAtBirth = tipStats[r.tipIdentity]
        if r.goneSession:
            goneIdentity = r.goneSession, r.location
            r.aliveTipsAtGone = tipStats[goneIdentity]
    tube.tipStats = tipStats

    logging.debug('Identified %s roots in sheet: %s' % (str(len(roots)), ws.title))
    logging.debug('Found %s roots in tube' % (str(len(tube.roots))))
    return tube


# tip stats code

def calculate_alive_tip_stats(tube):
    aliveTips = {}
    for existingRoot in tube.roots:
        tipID = existingRoot.tipIdentity
        if tipID not in aliveTips:
            aliveTips[tipID] = 1
        else:
            aliveTips[tipID] = aliveTips[tipID] + 1
    return aliveTips


def calculate_gone_tip_stats(din, tube):
    d = {}
    for i, c in sorted(din.iteritems()):
        for r in tube.roots:
            if (r.tipIdentity == i) and r.goneSession:
                gI = r.goneSession, r.location
                if gI not in d:
                    d[gI] = -1
                else:
                    d[gI] = d[gI] - 1
    return d


def tip_stats(tube):
    # get count of alive tips and the locations/sessions where roots die
    aliveStats = calculate_alive_tip_stats(tube)
    goneStats = calculate_gone_tip_stats(aliveStats, tube)
    # build a list of all root locations
    locations = []
    for r in tube.roots:
        if r.location not in locations:
            locations.append(r.location)
    # build a mapping of sessions to lists of locations
    loc_sessions = {}
    for L in locations:
        for k in aliveStats:
            sess, loc = k
            if L == loc:
                if L not in loc_sessions:
                    loc_sessions[L] = [sess]
                else:
                    loc_sessions[L].append(sess)
    for L in locations:
        for k in goneStats:
            sess, loc = k
            if L == loc:
                if L not in loc_sessions:
                    loc_sessions[L] = [sess]
                else:
                    loc_sessions[L].append(sess)
    # unique & sort the lists in loc_sessions
    for k in loc_sessions:
        i = list(set(loc_sessions[k]))
        i.sort()
        loc_sessions[k] = i
    # sum the alive and dead roots over time
    totalStats = {}
    for k in loc_sessions:
        temp = 0
        for sess in loc_sessions[k]:
            key = sess, k
            if key in aliveStats:
                temp = aliveStats[key] + temp
            if key in goneStats:
                temp = goneStats[key] + temp
            totalStats[key] = temp
    # return totalStats
    return totalStats


# root processing code

def root_from_row(row, indexDict):
    ''' build a root object from row, using indexDict for mappings'''
    # identification information
    rootName = row[indexDict['RootName']]
    location = row[indexDict['Location#']]
    birthSession = row[indexDict['BirthSession']]
    root = Root(rootName, location, birthSession)
    # check to see if anomalous root
    if row[indexDict['NumberOfTips']] == 1:
        root.anomaly = False
        if row[indexDict['TipLivStatus']].startswith('A'):
            root.isAlive = 'A'
        if row[indexDict['TipLivStatus']].startswith('G'):
            root.isAlive = 'G'
    else:
        root.anomaly = True
        root.isAlive = 'A'
    # root metadata
    root.session = row[indexDict['Session#']]
    if root.isAlive.startswith('G'):
        root.goneSession = root.session
    # order & diameter data
    root.order = row[indexDict['RootNotes']]
    root.avgDiameter = row[indexDict['TotAvgDiam(mm/10)']]
    return root


def get_tube_number(ws):
    ''' get tube number from a given worksheet.  this assumes that '''
    row1 = ws.rows[0]
    tubeIndex = get_index_by_value(row1, 'Tube#')
    tubeNumber = ws.columns[tubeIndex][1].value
    if tubeNumber:
        return tubeNumber
    else:
        return False


''' helper function for working with openpyxl'''


def get_root_sheets(wb):
    '''given a workbook, return all sheet names that end in "Root Synth""'''
    root_sheets = []
    sheets = wb.get_sheet_names()
    for sheet in sheets:
        if sheet.endswith('Root Synth'):
            root_sheets.append(sheet)
    return root_sheets


def stats(wb, sheetlist, value, verbose=False):
    '''generate counts of different values in a given column, based on column name
    this is done across a workbook'''
    dicty = {}
    for sheet in sheetlist:
        ws = wb.get_sheet_by_name(sheet)
        row1 = ws.rows[0]
        index = get_index_by_value(row1, value)
        for cell in ws.columns[index]:
            if cell.value in dicty:
                dicty[cell.value] = dicty[cell.value] + 1
            else:
                dicty[cell.value] = 1
                if verbose:
                    logging.info('Found new cell.value in sheet: %s' % (sheet))
    return dicty


def print_items_keys(iterable):
    for thing in iterable:
        for attr, value in thing.__dict__.iteritems():
            print attr, value
        print '================================'


# main function support
# http://code.activestate.com/recipes/577058/
def query_yes_no(question, default="yes"):
    """Ask a yes/no question via raw_input() and return their answer.
    "question" is a string that is presented to the user.
    "default" is the presumed answer if the user just hits <Enter>.
        It must be "yes" (the default), "no" or None (meaning
        an answer is required of the user).
        
    The "answer" return value is one of "yes" or "no".
    """
    valid = {"yes": True, "y": True, "ye": True,
             "no": False, "n": False}
    if default == None:
        prompt = " [y/n] "
    elif default == "yes":
        prompt = " [Y/n] "
    elif default == "no":
        prompt = " [y/N] "
    else:
        raise ValueError("invalid default answer: '%s'" % default)

    while True:
        sys.stdout.write(question + prompt)
        choice = raw_input().lower()
        if default is not None and choice == '':
            return valid[default]
        elif choice in valid:
            return valid[choice]
        else:
            sys.stdout.write("Please respond with 'yes' or 'no' " "(or 'y' or 'n').\n")


def write_out_data(output, tubes):
    '''
    write out the data from tubes in a xlsx workbook
    
    header data chosen by MSS
    '''
    # setup header indices
    header = ['RootName', 'Tube#', 'Location#', 'BirthSession', 'Birth Date',
              'Birth Year', 'GoneSession', 'Gone Date', 'Gone Year', 'Censored',
              'AliveTipsAtBirth', 'AliveTipsAtGone', 'Avg Diameter (mm/10)',
              'Order', 'Highest Order']
    headerIndex = {}
    # header index specifically for use with get_column_letter()
    for i in header:
        headerIndex[i] = header.index(i) + 1
    headerToRootMapping = {
        'RootName': 'rootName',
        'Location#': 'location',
        'BirthSession': 'birthSession',
        'GoneSession': 'goneSession',
        'Censored': 'censored',
        'AliveTipsAtBirth': 'aliveTipsAtBirth',
        'AliveTipsAtGone': 'aliveTipsAtGone',
        'Avg Diameter (mm/10)': 'avgDiameter',
        'Order': 'order',
        'Highest Order': 'highestOrder',
    }

    # setup workbook
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = 'Compiled Data'
    row_index = 1
    # write out header row
    for i in header:
        col_indx = headerIndex[i]
        col = openpyxl.cell.get_column_letter(col_indx)
        ws.cell('%s%s' % (col, row_index)).value = i
    row_index = row_index + 1
    # process the root data from each tube
    for tube in tubes:
        tubeNumber = tube.tubeNumber
        logging.debug('Writing out data for tube#: %s' % (str(tubeNumber)))
        for root in tube.roots:
            # write header
            col_indx = headerIndex['Tube#']
            col = openpyxl.cell.get_column_letter(col_indx)
            ws.cell('%s%s' % (col, row_index)).value = tubeNumber
            # build a dictionary of all the root items
            rootDict = root.__dict__
            for value in header:
                if value in headerToRootMapping:
                    rootValue = rootDict[headerToRootMapping[value]]
                    col_indx = headerIndex[value]
                    col = openpyxl.cell.get_column_letter(col_indx)
                    ws.cell('%s%s' % (col, row_index)).value = rootValue
                #
                # These two elif statements are not dynamic
                #
                elif value == 'Birth Date':
                    birthSession = rootDict['birthSession']
                    birthDate = tube.sessionDates[birthSession]
                    # specific to the data provided by MSS
                    birthYear = birthDate.split('.')[0]
                    col_indx = headerIndex[value]
                    col = openpyxl.cell.get_column_letter(col_indx)
                    ws.cell('%s%s' % (col, row_index)).value = birthDate
                    col_indx = headerIndex['Birth Year']
                    col = openpyxl.cell.get_column_letter(col_indx)
                    ws.cell('%s%s' % (col, row_index)).value = birthYear
                elif value == 'Gone Date':
                    # make sure root has kicked the bucket first.
                    if rootDict['censored'] == 0:
                        goneSession = rootDict['goneSession']
                        goneDate = tube.sessionDates[goneSession]
                        # specific to data provided by MSS
                        goneYear = goneDate.split('.')[0]
                        col_indx = headerIndex[value]
                        col = openpyxl.cell.get_column_letter(col_indx)
                        ws.cell('%s%s' % (col, row_index)).value = goneDate
                        col_indx = headerIndex['Gone Year']
                        col = openpyxl.cell.get_column_letter(col_indx)
                        ws.cell('%s%s' % (col, row_index)).value = goneYear
                elif value not in ['Birth Year', 'Gone Year', 'Tube#']:
                    logging.warning('Unhandled value in header: %s' % (value))
            row_index = row_index + 1
    # save results
    try:
        wb.save(filename=output)
    except Exception, e:
        logging.error('General error handler')
        logging.error('%s' % (e))
        return False
    return True


##
## Main program functions
##


def main(options):
    #
    #   DOES NOT HANDLE EXCEPTION LISTS
    #

    # file path verification and options handling
    if not options.verbose:
        logging.info('Output will not be verbose')
        logging.disable(logging.DEBUG)

    if not os.path.isfile(options.src_file):
        logging.error('specified source is not a file')
        sys.exit(-1)

    if os.path.exists(options.output):
        logging.warning('Specified output file already exists.\n')
        if not query_yes_no('Do you want to overwrite that file?', 'no'):
            logging.info('Exiting')
            sys.exit(-1)

    # open up src file
    try:
        wb = openpyxl.load_workbook(filename=options.src_file)
    except Exception, e:
        logging.error('general exception handler')
        logging.error('%s' % (e))
        sys.exit(-1)

    # get root data
    sheets = get_root_sheets(wb)
    if not sheets:
        logging.error('Could not find any WinRhizotron data sheets in the src file')
        logging.error('Make sure your workbook worksheets end in the string "Root Synth"')
        sys.exit(-1)

    tubes = []
    for sheet in sheets:
        try:
            ws = wb.get_sheet_by_name(sheet)
        except Exception, e:
            logging.error('Error occured attempting to get worksheet: %s' % (sheet))
            logging.error('%s' % (e))
            # attempt to get the next worksheet
            continue
        tube = process_worksheet(ws)
        if not tube:
            logging.error('Was not able to process the tube data from worksheet: %s' % (sheet))
        tubes.append(tube)
    if not tubes:
        logging.error('Was unable to get any tube data')
        sys.exit(-1)

    # write out the data for each tube into the new worksheet
    foo = write_out_data(options.output, tubes)
    if foo:
        logging.info('Sucesfully wrote out data to %s' % (options.output))
    else:
        logging.info('Did not write data out.')
        sys.exit(-1)

    logging.info('Done processing all data')
    sys.exit(0)


def root_options():
    parser = argparse.ArgumentParser(prog=__name__)
    parser.add_argument('-s', '--source', dest='src_file', required=True, type=str, action='store',
                        help='Source xlsx file to process')
    parser.add_argument('-e', '--exceptions', dest='exceptions', type=str, action='store', default=None,
                        help='A xlsx file containing root rewrite data for shifted roots')
    parser.add_argument('-o', '--output', dest='output', required=True, type=str, action='store',
                        help='Define the output file (xlsx)')
    parser.add_argument('-v', '--verbose', dest='verbose', default=False, action='store_true',
                        help='Enable verbose output')
    return parser


if __name__ == "__main__":
    parser = root_options()
    options = parser.parse_args()

    main(options)
